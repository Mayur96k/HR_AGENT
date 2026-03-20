"""
excel_exporter.py — Step 9: Export results to formatted Excel
FIX: New column sequence per requirement, adds Rank + Current Location + ATS Rating
"""
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# ── Column definition: (json_field, Excel header) in REQUIRED sequence ───────
COLUMNS = [
    ("applied_position",                "Applied Position"),
    ("name",                            "Name"),
    ("gender",                          "Gender"),
    ("current_location",                "Current Location"),
    ("total_working_experience",        "Total Working Experience"),
    ("relevance_experience",            "Relevance Experience"),
    ("graduation",                      "Graduation Degree"),
    ("post_graduate",                   "Post Graduation Degree"),
    ("current_company",                 "Current Company"),
    ("current_designation",             "Current Designation"),
    ("current_employment_tenure",       "Current Employment Tenure"),
    ("previous_employer",               "Previous Employer"),
    ("previous_designation",            "Previous Designation"),
    ("previous_employment_tenure",      "Previous Employment Tenure"),
    ("rank",                            "Rank"),
    ("ats_rating",                      "ATS Rating"),
    ("overall_fit_score",               "Fit Score (%)"),
    ("key_skills",                      "Key Skills"),
    ("recommendation",                  "Recommendation"),
    ("required_experience_match",       "Required Experience Match"),
    ("skill_match_summary",             "Skill Match Summary"),
    ("education_certifications_match",  "Education & Cert Match"),
    ("soft_skills_traits",              "Soft Skills Traits"),
    ("strengths",                       "Strengths"),
    ("concerns_or_gaps",                "Concerns / Gaps"),
    ("final_notes",                     "Final Notes"),
]

# Color bands per recommendation
REC_COLORS = {
    "highly recommended": "C6EFCE",   # green
    "recommended":        "EBFAEB",   # light green
    "maybe":              "FFEB9C",   # yellow
    "not recommended":    "FFC7CE",   # red
    "parse error":        "D9D9D9",   # grey
}

HDR_FILL    = PatternFill("solid", fgColor="1F3864")
HDR_FONT    = Font(bold=True, color="FFFFFF", size=11)
THIN_SIDE   = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
SCORE_FONT  = Font(bold=True, size=12)
RANK_FONT   = Font(bold=True, size=12, color="1F3864")


def _ats_rating(score: float) -> str:
    if score >= 80:
        return "Strong"
    if score >= 60:
        return "Moderate"
    return "Weak"


def export_to_excel(results: list, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "HR Evaluation"

    # Sort by overall_fit_score desc and assign rank
    results = sorted(
        results,
        key=lambda x: float(x.get("overall_fit_score", 0) or 0),
        reverse=True
    )
    for rank_idx, r in enumerate(results, 1):
        r["rank"] = rank_idx
        score = float(r.get("overall_fit_score", 0) or 0)
        r["ats_rating"] = _ats_rating(score)
        # Normalise score to 0-100
        r["overall_fit_score"] = round(score if score > 1 else score * 100)

    # ── Header row ────────────────────────────────────────────────────────────
    for col_idx, (_, header) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 32

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, r in enumerate(results, 2):
        rec       = str(r.get("recommendation", "")).lower()
        row_color = REC_COLORS.get(rec, "FFFFFF")
        fill      = PatternFill("solid", fgColor=row_color)

        for col_idx, (field, _) in enumerate(COLUMNS, 1):
            val  = r.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border    = THIN_BORDER

            if field == "overall_fit_score":
                cell.font      = SCORE_FONT
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif field == "rank":
                cell.font      = RANK_FONT
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif field == "ats_rating":
                cell.alignment = Alignment(horizontal="center", vertical="top")

    # ── Column widths (by position index) ────────────────────────────────────
    col_widths = {
        1: 22,   # Applied Position
        2: 22,   # Name
        3: 10,   # Gender
        4: 18,   # Current Location
        5: 20,   # Total Working Experience
        6: 20,   # Relevance Experience
        7: 20,   # Graduation Degree
        8: 22,   # Post Graduation Degree
        9: 24,   # Current Company
        10: 24,  # Current Designation
        11: 22,  # Current Employment Tenure
        12: 24,  # Previous Employer
        13: 24,  # Previous Designation
        14: 22,  # Previous Employment Tenure
        15: 8,   # Rank
        16: 12,  # ATS Rating
        17: 12,  # Fit Score (%)
        18: 38,  # Key Skills
        19: 20,  # Recommendation
        20: 22,  # Required Experience Match
        21: 36,  # Skill Match Summary
        22: 24,  # Education & Cert Match
        23: 26,  # Soft Skills Traits
        24: 36,  # Strengths
        25: 36,  # Concerns / Gaps
        26: 40,  # Final Notes
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Legend sheet ──────────────────────────────────────────────────────────
    ls = wb.create_sheet("Legend")
    ls.column_dimensions["A"].width = 22
    ls.column_dimensions["B"].width = 30
    ls.append(["Color", "Meaning"])
    ls.cell(1, 1).font = Font(bold=True)
    ls.cell(1, 2).font = Font(bold=True)

    for rec, color in REC_COLORS.items():
        row = ls.max_row + 1
        c = ls.cell(row=row, column=1, value=rec.title())
        c.fill = PatternFill("solid", fgColor=color)
        c.border = THIN_BORDER
        d = ls.cell(row=row, column=2, value=f"ATS: {'Strong' if 'highly' in rec else ('Moderate' if rec == 'recommended' else 'Weak')}")
        d.border = THIN_BORDER

    wb.save(output_path)
    log.info(f"Excel saved: {output_path}  ({len(results)} rows)")
